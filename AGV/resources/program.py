# Kurulum: 
#   pip install openpyxl

import openpyxl
from openpyxl import Workbook

# Yeni bir Excel dosyası oluştur
wb = Workbook()
ws = wb.active
ws.title = "Aylık Görev Listesi"

# Sütun başlıkları (ihtiyaca göre düzenleyebilirsiniz)
headers = ["Tarih", "Entegre 5", "Entegre 3", "N.Ş", "N.B"]
for col_num, header in enumerate(headers, start=1):
    ws.cell(row=1, column=col_num, value=header)

##############################################################################
# Fotoğraflardan okunabilen veriler
# (Eksik kısımları '?' ile gösteriliyor; elinizdeki net bilgilere göre doldurunuz)
##############################################################################

# 1) 10 Mart - 28 Mart (Pazartesi - Cuma)
#    Fotoğraflarda net şekilde sadece 10 Mart bilgisi görülüyor; diğer günler boş/eksik.

data_mart = [
    # 10 Mart (net görülenler)
    ["10/03/2025 Pazartesi", "Ömer - Selçuk", "Anıl - Berfin", "Ayra", "Taha"],
    
    # 11-14 Mart (fotoğrafta net bilgi yok, bu yüzden '?' veya '' konulabilir)
    ["11/03/2025 Salı",       "?", "?", "?", "?"],
    ["12/03/2025 Çarşamba",   "?", "?", "?", "?"],
    ["13/03/2025 Perşembe",   "?", "?", "?", "?"],
    ["14/03/2025 Cuma",       "?", "?", "?", "?"],

    # 17-21 Mart (fotoğrafta net bilgi seçilemiyor)
    ["17/03/2025 Pazartesi",  "?", "?", "?", "?"],
    ["18/03/2025 Salı",       "?", "?", "?", "?"],
    ["19/03/2025 Çarşamba",   "?", "?", "?", "?"],
    ["20/03/2025 Perşembe",   "?", "?", "?", "?"],
    ["21/03/2025 Cuma",       "?", "?", "?", "?"],

    # 24-28 Mart (fotoğrafta net bilgi seçilemiyor)
    ["24/03/2025 Pazartesi",  "?", "?", "?", "?"],
    ["25/03/2025 Salı",       "?", "?", "?", "?"],
    ["26/03/2025 Çarşamba",   "?", "?", "?", "?"],
    ["27/03/2025 Perşembe",   "?", "?", "?", "?"],
    ["28/03/2025 Cuma",       "?", "?", "?", "?"],
]

# 2) 31 Mart - 4 Nisan (Pazartesi - Cuma)
#    İkinci fotoğraftan okunan notlar. Net görülen kısımları tabloya ekledik.
#    Bazı yerler eksik olduğundan '?' kullandık.

data_nisan = [
    ["31/03/2025 Pazartesi", "?", "?", "Ayra - Taha",   "Berfin - Anıl"],
    ["01/04/2025 Salı",      "?", "?", "Ömer - Selçuk", "?"],
    ["02/04/2025 Çarşamba",  "?", "?", "İlayda - Taha", "?"],
    ["03/04/2025 Perşembe",  "?", "?", "Berfin - Ayra", "?"],
    ["04/04/2025 Cuma",      "?", "?", "Serife - Anıl", "?"],
]

##############################################################################
# Bütün veriyi tek listede birleştirelim
##############################################################################
data = data_mart + data_nisan

# Bu veriyi Excel'e yazalım
row_index = 2  # Başlık satırından sonraki satır
for row in data:
    for col_index, value in enumerate(row, start=1):
        ws.cell(row=row_index, column=col_index, value=value)
    row_index += 1

# Dosyayı kaydedelim
excel_dosya_adi = "gorev_listesi.xlsx"
wb.save(excel_dosya_adi)
print(f"Excel dosyası '{excel_dosya_adi}' oluşturuldu.")
